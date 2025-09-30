import os
import json
import datetime
import openpyxl
import argparse
import traceback
import numpy as np
from tqdm import tqdm
from collections import defaultdict
from openpyxl.styles import PatternFill, Font


def datetime_to_float(dt):
    excel_start_date = datetime.datetime(1899, 12, 30)
    delta = dt - excel_start_date
    return delta.days + delta.seconds / 86400.0


def transform_value(v):
    if isinstance(v, (int, float)):
        v = round(float(v), 2)
    elif isinstance(v, datetime.time):
        v = str(v)[:-3]
    elif isinstance(v, datetime.datetime):
        v = round(datetime_to_float(v), 0)
    elif isinstance(v, str):
        try:
            v = round(float(v), 2)
        except ValueError:
            pass
    return v.strip() if isinstance(v, str) else v


def compare_cell_value(v1, v2):

    v1 = transform_value(v1)
    v2 = transform_value(v2)
    if (v1 == "" and v2 is None) or (v1 is None and v2 == ""):
        return True
    if (v1 == "" and v2 == "") or (v1 is None and v2 is None):
        return True
    if type(v1) != type(v2):
        return False
    if v1 == v2:
        return True
    else:
        return False


def compare_fill_color(fill1, fill2):
    fgColor1 = fill1.fgColor.rgb if fill1.fgColor else None
    fgColor2 = fill2.fgColor.rgb if fill2.fgColor else None
    bgColor1 = fill1.bgColor.rgb if fill1.bgColor else None
    bgColor2 = fill2.bgColor.rgb if fill2.bgColor else None

    if fgColor1 == fgColor2 and bgColor1 == bgColor2:
        return True
    else:
        return False


def compare_font_color(font_gt, font_proc):
    if font_gt.color is not None and font_proc.color is not None:
        return font_gt.color.rgb == font_proc.color.rgb
    elif font_gt.color is None and font_proc.color is None:
        return True
    else:
        return False


def col_num2name(n):
    """ Convert a column number to an Excel column name """
    name = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        name = chr(65 + remainder) + name
    return name


def col_name2num(name):
    """ Convert an Excel column name to a column number """
    num = 0
    for c in name:
        num = num * 26 + (ord(c) - ord('A') + 1)
    return num


def parse_cell_range(range_str):
    """ Parse a range string like 'A1:AB12' """
    start_cell, end_cell = range_str.split(':')
    start_col, start_row = '', ''
    for char in start_cell:
        if char.isdigit():
            start_row += char
        else:
            start_col += char
    
    end_col, end_row = '', ''
    for char in end_cell:
        if char.isdigit():
            end_row += char
        else:
            end_col += char

    return (col_name2num(start_col), int(start_row)), (col_name2num(end_col), int(end_row))


def generate_cell_names(range_str):
    """ Generate a list of all cell names in the specified range """
    if ':' not in range_str:
        return [range_str]
    (start_col, start_row), (end_col, end_row) = parse_cell_range(range_str)
    columns = [col_num2name(i) for i in range(start_col, end_col + 1)]
    cell_names = [f"{col}{row}" for col in columns for row in range(start_row, end_row + 1)]
    return cell_names


def cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range, is_CF):
    if sheet_name not in wb_proc:
        return False, "worksheet not found"
    ws_gt = wb_gt[sheet_name]
    ws_proc = wb_proc[sheet_name]

    cell_names = generate_cell_names(cell_range)

    for cell_name in cell_names:
        cell_gt = ws_gt[cell_name]
        cell_proc = ws_proc[cell_name]

        if not compare_cell_value(cell_gt.value, cell_proc.value):
            msg = f"Value difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.value},\
                    ws_proc has {cell_proc.value}"
            return False, msg
        
        if is_CF:
            if not compare_fill_color(cell_gt.fill, cell_proc.fill):
                msg = f"Fill color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.fill.fgColor.rgb},\
                        ws_proc has {cell_proc.fill.fgColor.rgb}"
                return False, msg

            if not compare_font_color(cell_gt.font, cell_proc.font):
                msg = f"Font color difference at cell {cell_gt.coordinate}: ws_gt has {cell_gt.font.color.rgb},\
                        ws_proc has {cell_proc.font.color.rgb}"
                return False, msg

    return True, "Success"


def compare_workbooks(gt_file, proc_file, instruction_type, answer_position):
    if not os.path.exists(proc_file):
        return False, "File not exist"
    # Open workbooks
    if "CF" in proc_file:
        is_CF = True
    else:
        is_CF = False
    try:
        # just_open(wb_proc)
        wb_gt = openpyxl.load_workbook(filename=gt_file, data_only=True)
        wb_proc = openpyxl.load_workbook(filename=proc_file, data_only=True)
    except Exception as e:
        return False, str(e)

    # Initialize report
    result = False
    msg = ""

    sheet_cell_ranges = answer_position.split(',')
    for sheet_cell_range in sheet_cell_ranges:
        if '!' in sheet_cell_range:
            sheet_name, cell_range = sheet_cell_range.split('!')
            sheet_name = sheet_name.lstrip("'").rstrip("'")
        else:
            sheet_name = wb_gt.sheetnames[0]
            cell_range = sheet_cell_range
    result, msg = cell_level_compare(wb_gt, wb_proc, sheet_name, cell_range , is_CF)

    return result, msg

def run_solution_one_data(data, client, runs_name):
    data['test_case_results'] = []
    data['test_case_messages'] = []

    for idx in range(1, 4):
        try:
            input_file = f"{idx}_{data['id']}_input.xlsx"
            output_file = f"{idx}_{data['id']}_output.xlsx"

            if data['solution']:
                local_solution = data['solution'].replace(f"1_{data['id']}_input.xlsx", input_file)
                local_solution = local_solution.replace(f"1_{data['id']}_output.xlsx", output_file)
                client.execute(local_solution)

            gt_file = f"data/spreadsheet/{data['id']}/{idx}_{data['id']}_answer.xlsx"
            output_file = f"outs/{runs_name}/spreadsheet/{idx}_{data['id']}_output.xlsx"
        
            r, message = compare_workbooks(gt_file, output_file, data['instruction_type'], data['answer_position'])
        except:
            r, message = 0, traceback.format_exc()

        data['test_case_results'].append(int(r))
        data['test_case_messages'].append(message)
    
    data['total_soft_restriction'] = sum(data['test_case_results']) / len(data['test_case_results'])
    data['total_hard_restriction'] = 1.0 if all(data['test_case_results']) else 0.0

    return data

if __name__ == "__main__":
    r, message = compare_workbooks(
        '/mnt/cache/data/SpreadsheetBench/data/spreadsheet/283-32/1_283-32_answer.xlsx', 
        '/mnt/cache/data/SpreadsheetBench/outs/yaml_desc/spreadsheet/1_283-32_output.xlsx', 
        "Sheet-Level Manipulation", 
        "Sheet3'!A:G,'Sheet4'!A:G"
    )

    print(message)