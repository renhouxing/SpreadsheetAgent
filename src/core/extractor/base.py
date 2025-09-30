import os
import json
import traceback

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from ..client import ClientJupyterKernel

EXTRACTOR = {}

def register(name):
    def decorator(cls):
        EXTRACTOR[name] = cls
        return cls
    return decorator

repo_dir = os.path.abspath(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

class StructureExtractor:

    def __init__(self, *args, **kwargs):

        self.url = kwargs.get('url', 'ocalhost:8000')
        self.vision_url = kwargs.get('vision_url', 'localhost:8000')

        self.excel2image_url = kwargs.get('excel2image_url', 'localhost:8007')
        self.code_exec_url = kwargs.get('code_exec_url', 'localhost:8081')

        self.suffix = kwargs.get('suffix', 'default')

        self.force = kwargs.get('force', False)

        self.model_params = {
            'top_p': kwargs.get('top_p', 0.95),
            'temperature': kwargs.get('temperature', 0.6)
        }

    def get_used_ranges(self, file_path):
        try:
            wb = load_workbook(file_path, read_only=True, data_only=True)
            used_ranges = {}

            for ws in wb.worksheets:
                min_row = ws.min_row
                max_row = ws.max_row
                min_col = ws.min_column
                max_col = ws.max_column

                if min_row is None:
                    used_ranges[ws.title] = None
                else:
                    used_ranges[ws.title] = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
        except:
            used_ranges = {}

        return used_ranges
    
    def get_structure_sheet(self, docker_path, sheet_name, used_range, client, mount_dir):
        pass
    
    def __call__(self, data):
        docker_path = f"/mnt/data/{data['input_file']}"
        client = ClientJupyterKernel(self.code_exec_url, mount_dir=data['mount_dir'])

        for sheet_name, used_range in self.get_used_ranges(os.path.join(data['real_dir'], data['input_file'])).items():
            target_path = os.path.join(data['real_dir'], data['input_file'].replace('.xlsx', f'_{sheet_name}_{self.suffix}.json'))

            if os.path.exists(target_path) and not self.force:
                continue

            for _ in range(3):
                messages, structure, error, other = self.get_structure_sheet(
                    docker_path, sheet_name, used_range, client, data['mount_dir']
                )

                if structure:
                    with open(target_path, 'w') as f:
                        other.update(dict(
                            messages=messages, 
                            structure=structure, 
                            error=error
                        ))
                        json.dump(other, f, indent=4, ensure_ascii=False)
                
                if error is None:
                    break