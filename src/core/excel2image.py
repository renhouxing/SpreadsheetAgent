from flask import Flask, request, jsonify
import tempfile, os, base64
import excel2img
import traceback
app = Flask(__name__)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "没有发现文件字段"}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "没有选择文件"}), 400

    if file:
        file.save(file.filename)
        return jsonify({"message": "文件上传成功", "filename": file.filename}), 200
    else:
        return jsonify({"error": "文件类型不允许"}), 400

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter 

def used_range(file, sheet):
    wb = load_workbook(file, data_only=True)
    ws = wb[sheet]
    
    return f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

@app.route("/excel2img", methods=["POST"])
def excel_to_img():
    if "file" not in request.files:
        return jsonify({"error": "missing file"}), 400
    
    excel_file = request.files["file"]

    with tempfile.TemporaryDirectory() as tmpdir:
        in_path = os.path.join(tmpdir, excel_file.filename or "input.xlsx")
        out_path = os.path.join(tmpdir, "out.png")
        excel_file.save(in_path)

        page = request.form.get('page')
        _range = request.form.get('_range', None)

        if _range is None:
            _range = used_range(in_path, page)

        try:
            excel2img.export_img(in_path, out_path, page, _range)
        except:
            return jsonify({"error": traceback.format_exc()}), 500

        if not os.path.exists(out_path):
            return jsonify({"error": "save file failed"}), 500

        with open(out_path, "rb") as f:
            b64 = base64.b64encode(f.read()).decode("utf-8")

    return jsonify({
        "image_base64": b64
    })
    
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8007, debug=True)