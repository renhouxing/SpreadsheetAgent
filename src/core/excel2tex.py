from __future__ import annotations
from typing import Tuple, Dict, Optional, List
from collections import defaultdict

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.cell import range_boundaries

# ----------------------------
# LaTeX helpers
# ----------------------------

LATEX_SPECIALS = {
    "&": r"\&", "%": r"\%", "$": r"\$", "#": r"\#",
    "_": r"\_", "{": r"\{", "}": r"\}", "~": r"\textasciitilde{}",
    "^": r"\textasciicircum{}", "\\": r"\textbackslash{}",
}

def latex_escape(s: str, enabled: bool) -> str:
    if not enabled:
        return s
    return "".join(LATEX_SPECIALS.get(ch, ch) for ch in s)

# ----------------------------
# Merge handling and line-planning
# ----------------------------

def build_merge_maps(ws: Worksheet, sel_box: Tuple[int,int,int,int]):
    """
    Returns:
      anchor_of[(r,c)] -> (vspan, hspan) for merges (clipped to selection)
      covered[(r,c)] -> True if cell is covered by a merge (not the anchor)
      placeholders_by_row[r] -> list of (c_start, hspan) placeholders to emit in later rows
      vspans_by_row_start[r] -> list of (c1, c2, remaining_rows_below_anchor)
                                for drawing horizontal rules that avoid multirows
    All coordinates are absolute sheet (1-based rows/cols).
    """
    (sr1, sc1, sr2, sc2) = sel_box
    anchor_of: Dict[Tuple[int,int], Tuple[int,int]] = {}
    covered: Dict[Tuple[int,int], bool] = {}
    placeholders_by_row: Dict[int, List[Tuple[int,int]]] = defaultdict(list)
    vspans_by_row_start: Dict[int, List[Tuple[int,int,int]]] = defaultdict(list)

    for m in ws.merged_cells.ranges:
        mr1, mc1, mr2, mc2 = m.min_row, m.min_col, m.max_row, m.max_col
        # intersect merge rect with selection rect
        ir1 = max(mr1, sr1); ic1 = max(mc1, sc1)
        ir2 = min(mr2, sr2); ic2 = min(mc2, sc2)
        if ir1 > ir2 or ic1 > ic2:
            continue

        vspan = ir2 - ir1 + 1
        hspan = ic2 - ic1 + 1
        anchor = (ir1, ic1)
        anchor_of[anchor] = (vspan, hspan)

        # mark covered cells (except anchor)
        for r in range(ir1, ir2 + 1):
            for c in range(ic1, ic2 + 1):
                if (r, c) != anchor:
                    covered[(r, c)] = True

        # placeholders in rows below the anchor to keep column count & vertical lines
        for r in range(ir1 + 1, ir1 + vspan):
            placeholders_by_row[r].append((ic1, hspan))

        # track a vertical span that blocks horizontal lines for next (vspan-1) boundaries
        if vspan > 1:
            vspans_by_row_start[ir1].append((ic1, ic2, vspan - 1))

    return anchor_of, covered, placeholders_by_row, vspans_by_row_start

def cline_segments_for_row(n_cols: int, active_blocked: List[Tuple[int,int]]) -> List[Tuple[int,int]]:
    """
    Given a list of blocked closed intervals [c1,c2] (1-based col indices) where a horizontal rule
    must NOT cross (because a multirow continues), return drawable segments [i,j] for \cline/\cmidrule.
    """
    if not active_blocked:
        return [(1, n_cols)]
    blocked = [False] * (n_cols + 1)  # index 1..n_cols
    for c1, c2 in active_blocked:
        c1 = max(1, c1); c2 = min(n_cols, c2)
        for k in range(c1, c2 + 1):
            blocked[k] = True

    segs = []
    k = 1
    while k <= n_cols:
        while k <= n_cols and blocked[k]:
            k += 1
        if k > n_cols:
            break
        start = k
        while k <= n_cols and not blocked[k]:
            k += 1
        end = k - 1
        segs.append((start, end))
    return segs

# ----------------------------
# Column spec helpers
# ----------------------------

def build_column_format(n_cols: int, column_format: Optional[str], vertical_lines: bool) -> str:
    """
    If user supplies column_format, use it as-is.
    Else: when vertical_lines=True, default to |c|c|...| ; otherwise 'c'*n.
    """
    if column_format:
        return column_format
    if vertical_lines:
        return "|" + "|".join("c" for _ in range(n_cols)) + "|"
    return "c" * n_cols

def multicol_spec(vertical_lines: bool) -> str:
    """
    Build the 2nd arg of \multicolumn. To keep vertical rules consistent when vertical_lines=True,
    we simply use {|c|}. When vertical_lines=False, we use {c}.
    """
    return "|c|" if vertical_lines else "c"

# ----------------------------
# Main converter
# ----------------------------

def excel_range_to_latex(
    xlsx_path: str,
    sheet: str,
    cell_range: str,
    *,
    escape_text: bool = True,
    vertical_lines: bool = True,   # default: show vertical column lines
    booktabs: bool = False,        # set True only when vertical_lines=False (booktabs discourages vertical lines)
    column_format: Optional[str] = None,
) -> str:
    """
    Convert a rectangular Excel selection (sheet+range) to LaTeX WITHOUT any colors.

    - vertical_lines=True -> table uses |c|c|...| and \hline/\cline, with correct placeholders.
    - vertical_lines=False + booktabs=True -> booktabs style (\toprule/\midrule/\cmidrule/\bottomrule).
    - merged cells supported with \multirow + \multicolumn; later rows get multicolumn empty placeholders.
    """
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
    ws = wb[sheet]

    # openpyxl gives (min_col, min_row, max_col, max_row)
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    sr1, sc1, sr2, sc2 = min_row, min_col, max_row, max_col
    n_cols = sc2 - sc1 + 1

    sel_box = (sr1, sc1, sr2, sc2)
    anchor_of, covered, placeholders_by_row, vspans_by_row_start = build_merge_maps(ws, sel_box)

    tab_align = build_column_format(n_cols, column_format, vertical_lines)
    out_lines: List[str] = []
    out_lines.append(r"\begin{tabular}{" + tab_align + "}")

    # top rule
    if booktabs and not vertical_lines:
        out_lines.append(r"\toprule")
    else:
        out_lines.append(r"\hline")

    # Track vertical spans that continue below each printed row boundary
    # Each entry: (c1, c2, remaining_rows_below_current_boundary) in 1..n_cols coords
    active_vspans: List[Tuple[int,int,int]] = []

    spec = multicol_spec(vertical_lines)

    for r in range(sr1, sr2 + 1):
        # add any vspans starting at this row (convert columns to 1..n_cols)
        for (c1_abs, c2_abs, rem) in vspans_by_row_start.get(r, []):
            c1_rel = c1_abs - sc1 + 1
            c2_rel = c2_abs - sc1 + 1
            active_vspans.append((c1_rel, c2_rel, rem))

        # collect cells for this row
        row_cells: List[str] = []
        c = sc1
        row_phs = sorted(placeholders_by_row.get(r, []))  # list[(c_start_abs, hspan)]
        ph_idx = 0

        while c <= sc2:
            # Placeholder due to a vertical span above
            if ph_idx < len(row_phs) and c == row_phs[ph_idx][0]:
                hspan = row_phs[ph_idx][1]
                row_cells.append(rf"\multicolumn{{{hspan}}}{{{spec}}}{{}}")
                c += hspan
                ph_idx += 1
                continue

            # Covered by some merge (same row) -> skip; anchor will handle it
            if (r, c) in covered:
                c += 1
                continue

            # Visible cell (maybe anchor)
            val = ws.cell(row=r, column=c).value
            s = "" if val is None else str(val)
            s = latex_escape(s, enabled=escape_text)

            vspan, hspan = 1, 1
            if (r, c) in anchor_of:
                vspan, hspan = anchor_of[(r, c)]

            if vspan > 1 and hspan > 1:
                # Anchor of a vertical+horizontal merge: we emit multicolumn around multirow.
                cell_tex = rf"\multicolumn{{{hspan}}}{{{spec}}}{{\multirow{{{vspan}}}{{*}}{{{s}}}}}"
            elif vspan > 1:
                cell_tex = rf"\multirow{{{vspan}}}{{*}}{{{s}}}"
            elif hspan > 1:
                cell_tex = rf"\multicolumn{{{hspan}}}{{{spec}}}{{{s}}}"
            else:
                cell_tex = s

            row_cells.append(cell_tex)
            c += hspan

        out_lines.append(" & ".join(row_cells) + r" \\")

        # horizontal rule below this row
        blocked_ranges = [(c1, c2) for (c1, c2, rem) in active_vspans if rem > 0]
        segs = cline_segments_for_row(n_cols, blocked_ranges)

        if booktabs and not vertical_lines:
            # booktabs: prefer \midrule when no blocks; otherwise segmented \cmidrule
            if not blocked_ranges:
                out_lines.append(r"\midrule")
            else:
                for i, j in segs:
                    out_lines.append(rf"\cmidrule(lr){{{i}-{j}}}")
        else:
            # vertical lines style: \hline if contiguous; else segmented \cline
            if not blocked_ranges:
                out_lines.append(r"\hline")
            else:
                for i, j in segs:
                    out_lines.append(rf"\cline{{{i}-{j}}}")

        # advance vertical spans (we just crossed one boundary)
        new_active: List[Tuple[int,int,int]] = []
        for (c1, c2, rem) in active_vspans:
            if rem - 1 > 0:
                new_active.append((c1, c2, rem - 1))
        active_vspans = new_active

    # bottom rule for booktabs if used (and no vertical lines)
    if booktabs and not vertical_lines:
        out_lines[-1] = r"\bottomrule"

    out_lines.append(r"\end{tabular}")
    return "\n".join(out_lines)

def convert_excel_to_latex(path: str, sheet_name: str, _range: str):
    try:
        return excel_range_to_latex(path, sheet_name, _range)
    except Exception as e:
        return f"Error converting Excel to LaTeX: {e}"

if __name__ == "__main__":
    # Example usage
    path = "/mnt/cache/data/SpreadsheetBench/data/spreadsheet/5-33/1_5-33_input.xlsx"
    sheet = "Count AD"
    rng = "A1:L92"
    tex = convert_excel_to_latex(
        path=path,
        sheet_name=sheet,
        _range=rng,
    )
    print(tex)
